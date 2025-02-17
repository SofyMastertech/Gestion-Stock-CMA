from PySide6.QtWidgets import QFileDialog, QMessageBox, QDialog, QVBoxLayout, QLabel, QRadioButton, QDialogButtonBox
from PySide6.QtCore import QThread, Signal
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageTemplate, Frame
from reportlab.lib.styles import getSampleStyleSheet
from datetime import datetime
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP
from collections import defaultdict

#################################################
#                THREAD ET DIALOGUES            #
#################################################

class ExportThread(QThread):
    """
    Thread pour exécuter l'exportation en arrière-plan.
    """
    finished = Signal(bool, str)

    def __init__(self, export_func, filename, headers, data, export_mode="individual"):
        super().__init__()
        self.export_func = export_func
        self.filename = filename
        self.headers = headers
        self.data = data
        self.export_mode = export_mode

    def run(self):
        try:
            self.export_func(self.filename, self.headers, self.data, self.export_mode)
            self.finished.emit(True, "Exportation réussie")
        except Exception as e:
            self.finished.emit(False, str(e))


class ExportOptionsDialog(QDialog):
    """
    Boîte de dialogue pour choisir le mode d'exportation (individuel ou moyenne).
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Options d'exportation")
        self.setMinimumWidth(300)
        layout = QVBoxLayout(self)

        self.individual_radio = QRadioButton("Exporter les valeurs individuelles")
        self.average_radio = QRadioButton("Exporter les moyennes pour tests multi-lots")
        self.individual_radio.setChecked(True)

        layout.addWidget(QLabel("Choisissez le mode d'exportation :"))
        layout.addWidget(self.individual_radio)
        layout.addWidget(self.average_radio)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(self.accept)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def get_export_mode(self):
        return "individual" if self.individual_radio.isChecked() else "average"


def export_data(parent, export_type, headers, data):
    """
    Ouvre une boîte de dialogue pour choisir le fichier à exporter et lance le processus.
    """
    dialog = ExportOptionsDialog()
    if dialog.exec_() == QDialog.Accepted:
        export_mode = dialog.get_export_mode()

        # Filtre de fichier
        if export_type == "excel":
            file_filter = "Fichiers Excel (*.xlsx)"
            default_extension = ".xlsx"
        elif export_type == "pdf":
            file_filter = "Fichiers PDF (*.pdf)"
            default_extension = ".pdf"
        else:
            file_filter = f"Fichiers {export_type.upper()} (*.{export_type.lower()})"
            default_extension = f".{export_type.lower()}"

        # Choisir l'emplacement
        filename, _ = QFileDialog.getSaveFileName(parent, "Exporter", "", file_filter)
        if not filename:
            return

        # Vérifier l'extension
        if not filename.lower().endswith(default_extension):
            filename += default_extension

        # Choisir la bonne fonction
        export_func = export_to_pdf if export_type == "pdf" else export_to_excel

        # Lancer l'export dans un thread
        parent.export_thread = ExportThread(export_func, filename, headers, data, export_mode)
        parent.export_thread.finished.connect(parent.on_export_finished)
        parent.export_thread.start()

#################################################
#               EXPORTATION PDF                 #
#################################################

class MyDocTemplate(SimpleDocTemplate):
    """
    Modèle personnalisé pour PDF en paysage.
    """
    def __init__(self, filename, **kw):
        kw['pagesize'] = landscape(A4)
        super().__init__(filename, **kw)
        margin = 20
        self.page_width, self.page_height = kw['pagesize']

        template = PageTemplate(
            'normal',
            [
                Frame(
                    margin,
                    margin,
                    self.page_width - (2 * margin),
                    self.page_height - (2 * margin),
                    id='normal'
                ),
            ]
        )
        self.addPageTemplates([template])

    def afterPage(self):
        """
        Pied de page personnalisé.
        """
        canvas = self.canv
        margin = 20
        canvas.saveState()

        # Ligne séparatrice
        canvas.line(margin, margin + 5, self.page_width - margin, margin + 5)

        # Footer
        canvas.setFont('Helvetica', 8)
        canvas.drawString(margin, margin / 2, "Laboratoire ELRYM - Tous droits réservés")
        page_num = canvas.getPageNumber()
        canvas.drawRightString(self.page_width - margin, margin / 2, f"Page {page_num}")

        canvas.restoreState()


def export_to_pdf(filename, headers, data, export_mode="individual"):
    """
    Exporte les données en PDF avec ou sans moyenne.
    """
    doc = MyDocTemplate(filename)
    elements = []

    styles = getSampleStyleSheet()
    title_style = deepcopy(styles["Title"])
    title_style.alignment = 1

    # Titre
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Laboratoire ELRYM (Analyses Médicales)", title_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph("Rapport d'Exportation", styles["Heading2"]))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 20))

    # Calcul des moyennes si nécessaire
    if export_mode == "average":
        headers, data = calculate_averages(headers, data)

    # Largeur des colonnes
    page_width, _ = landscape(A4)
    margin = 20
    available_width = page_width - 2 * margin
    col_widths = [available_width / len(headers)] * len(headers)

    # Style du tableau
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
    ])

    # Construction du tableau
    data_table = [headers] + data
    table = Table(data_table, colWidths=col_widths, repeatRows=1)
    table.setStyle(table_style)
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Laboratoire ELRYM - Tous droits réservés", styles["Normal"]))

    doc.build(elements)

#################################################
#               EXPORTATION EXCEL               #
#################################################

def export_to_excel(filename, headers, data, export_mode="individual"):
    """
    Exporte les données dans un fichier Excel, avec ou sans moyenne.
    """
    try:
        # === Workbook ===
        wb = Workbook()
        ws = wb.active
        ws.title = "Données Exportées"

        # === Styles généraux ===
        title_font = Font(size=16, bold=True, color="FFFFFF")
        subtitle_font = Font(size=12, italic=True, color="FFFFFF")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        data_font = Font(size=10)
        footer_font = Font(size=9, italic=True, color="555555")

        header_fill = PatternFill(start_color="4A90E2", end_color="357ABD", fill_type="solid")
        alt_row_fill = PatternFill(start_color="EAF1FB", end_color="EAF1FB", fill_type="solid")
        border_style = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        center_align = Alignment(horizontal="center", vertical="center")

        # === Titre ===
        ws.merge_cells("A1:{}1".format(get_column_letter(len(headers))))
        title_cell = ws["A1"]
        title_cell.value = "Rapport d'Exportation - Laboratoire ELRYM"
        title_cell.font = title_font
        title_cell.alignment = center_align
        title_cell.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")

        # === Sous-titre (date) ===
        ws.merge_cells("A2:{}2".format(get_column_letter(len(headers))))
        subtitle_cell = ws["A2"]
        subtitle_cell.value = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        subtitle_cell.font = subtitle_font
        subtitle_cell.alignment = center_align
        subtitle_cell.fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")

        # Ligne vide
        ws.append([])

        # === Si mode moyenne, calculer d'abord les moyennes ===
        if export_mode == "average":
            headers, data = calculate_averages(headers, data)

        # === En-têtes ===
        ws.append(headers)
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f"{col_letter}4"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border_style

        # === Insertion des données ===
        row_start_index = 5
        for row_idx, row_values in enumerate(data, start=row_start_index):
            for col_idx in range(len(headers)):
                col_letter = get_column_letter(col_idx + 1)
                cell = ws[f"{col_letter}{row_idx}"]
                if col_idx < len(row_values):
                    cell.value = row_values[col_idx]
                else:
                    cell.value = "N/A"
                cell.font = data_font
                cell.alignment = center_align
                cell.border = border_style
                if row_idx % 2 == 0:
                    cell.fill = alt_row_fill

        # === Ajustement automatique des colonnes ===
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].auto_size = True

        # === Pied de page ===
        footer_row = ws.max_row + 2
        ws.merge_cells(f"A{footer_row}:{get_column_letter(len(headers))}{footer_row}")
        footer_cell = ws[f"A{footer_row}"]
        footer_cell.value = "🔬 Laboratoire ELRYM - Tous droits réservés | Contact: info@elrym.com"
        footer_cell.font = footer_font
        footer_cell.alignment = center_align
        footer_cell.fill = PatternFill(start_color="E3E3E3", end_color="E3E3E3", fill_type="solid")

        # Sauvegarde
        wb.save(filename)
        print(f"Le fichier '{filename}' a été créé avec succès.")

    except Exception as e:
        print(f"Une erreur s'est produite lors de l'exportation : {e}")

#################################################
#     NORMALISATION & CALCULS DE MOYENNES       #
#################################################

def normalize_analyte_names(headers, data):
    """
    Convertit les noms d'analytes en majuscules pour éviter les doublons.
    """
    try:
        analyte_index = headers.index("Nom analyte")
    except ValueError:
        return headers, data

    normalized_data = []
    for row in data:
        row_copy = list(row)
        row_copy[analyte_index] = row_copy[analyte_index].upper()
        normalized_data.append(row_copy)

    return headers, normalized_data


def calculate_averages(headers, data):
    """
    Calcule les moyennes des colonnes numériques pour chaque analyte.
    - Arrondi à une décimale pour obtenir exactement 92.6 au lieu de 92.7 ou 92.66.
    - Si l'analyte n'a qu'un seul lot, on garde la ligne telle quelle.
    - Sinon, on crée une ligne unique contenant la moyenne de toutes les colonnes numériques identifiées.
    - Pour les colonnes 'Date Ouverture', 'Date Fin' et 'Durée', on met 'MSPL' dans la ligne de moyenne.
    """
    try:
        analyte_index = headers.index("Nom analyte")
    except ValueError:
        return headers, data  # Pas de colonne 'Nom analyte', on renvoie tel quel

    # Déterminer les colonnes numériques en fonction des en-têtes
    if "Perte (%)" in headers:  # Mode Volume par test
        numeric_candidates = [
            "Tests Estimés", "Tests Réalisés", "Perte (Tests)", "Facteur Utilisation", "Perte (%)"
        ]
    else:  # Mode Tests
        numeric_candidates = [
            "Volume Total (ml)", "Volume Restant (ml)", "Tests Réalisés", "Volume/Test (ml)", "Perte %"
        ]

    # On récupère leurs indices si elles existent
    numeric_columns = []
    for candidate in numeric_candidates:
        if candidate in headers:
            numeric_columns.append(headers.index(candidate))

    # Grouper les lignes par analyte (insensible à la casse)
    groups = defaultdict(list)
    for row in data:
        analyte_name = str(row[analyte_index]).strip().upper()
        groups[analyte_name].append(row)

    averaged_rows = []
    for analyte, rows_group in groups.items():
        # S'il n'y a qu'un seul lot pour cet analyte, on ne fait pas de moyenne
        if len(rows_group) == 1:
            averaged_rows.append(rows_group[0])
            continue

        # Créer la ligne de moyenne
        avg_row = [""] * len(headers)
        avg_row[analyte_index] = analyte.capitalize()  # Nom analyte en capitalisant la 1re lettre

        # Calculer la moyenne ARRONDI à une décimale
        for col_idx in numeric_columns:
            numeric_values = []
            for row_data in rows_group:
                val = row_data[col_idx]
                try:
                    # Nettoyer la valeur s'il s'agit d'une chaîne
                    if isinstance(val, str):
                        val = val.strip()
                    dec_val = Decimal(str(val))
                    numeric_values.append(dec_val)
                except (ValueError, TypeError):
                    pass

            if numeric_values:
                total = sum(numeric_values)
                mean_val = total / len(numeric_values)
                # Arrondi à 1 décimale (92.666… => 92.7)
                mean_val = mean_val.quantize(Decimal("0.0"), rounding=ROUND_HALF_UP)
                avg_row[col_idx] = str(mean_val)
            else:
                avg_row[col_idx] = "N/A"

        # Pour les autres colonnes, logique PDF : "Date Ouverture", "Date Fin", "Durée" => "MSPL"
        for col_idx in range(len(headers)):
            if col_idx not in numeric_columns and col_idx != analyte_index:
                if headers[col_idx] in ["Date Ouverture", "Date Fin", "Durée"]:
                    avg_row[col_idx] = "MSPL"
                else:
                    # Conserver la première valeur
                    avg_row[col_idx] = rows_group[0][col_idx]

        averaged_rows.append(avg_row)

    return headers, averaged_rows